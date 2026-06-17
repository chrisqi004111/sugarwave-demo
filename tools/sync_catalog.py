#!/usr/bin/env python
"""Sync the Scene Lab product catalogue.

Source of truth split:
  - which products exist + their CATEGORY  -> the folder layout
        src/assets/catalog/<category>/<id>.png
  - name / price / dimensions / tags       -> product-catalog.xlsx (human-edited)

Running this script:
  1. scans the catalog folders for {id: category}
  2. reads product-catalog.xlsx (if present) for the editable fields
  3. writes src/assets/catalog/meta.json   (consumed by TrialPage.jsx)
  4. rewrites product-catalog.xlsx so it always lists every current product
     (new images get a blank row to fill in; deleted images drop out)

Usage:  python tools/sync_catalog.py
"""
import json
import os
import re
import glob
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CATALOG_DIR = os.path.join(ROOT, 'src', 'assets', 'catalog')
META_JSON = os.path.join(CATALOG_DIR, 'meta.json')
XLSX = os.path.join(ROOT, 'product-catalog.xlsx')

COLUMNS = ['id', 'category', 'name', 'price', 'dimensions', 'tags', 'ai_pick']
IMG_EXT = ('.png', '.jpg', '.jpeg', '.webp')
TRUTHY = {'1', 'yes', 'y', 'true', 't', '✓', 'x'}

# Seed values — used only to pre-fill a product the first time it appears
# (before anyone has typed real numbers into the Excel). Edit the Excel, not this.
# ai_pick = 'yes' lets the product show up in the preset AI Recommendation.
# Keep it to neutral-toned pieces (transparent / black-white / marble); the demo
# scenes are cream / earth-toned, so vibrant products read as 'no'.
# Values below pulled from sugarwave_price_list_2026.xlsx (中英文 sheet,
# "Price (EUR)" + "Size" columns). starberry / strawberrys have no matching
# row in that list yet — left as placeholders (confirm which product they are).
SEED = {
    'boop':        {'name': 'Boop', 'price': 180, 'dimensions': '18cm × 18cm × 16cm', 'tags': 'minimal, modern', 'ai_pick': 'yes'},
    'fluffy':      {'name': 'Fluffy', 'price': 70, 'dimensions': '19cm × 15.5cm × 12cm', 'tags': 'organic, modern', 'ai_pick': 'no'},
    'calla-amber': {'name': 'Calla — Amber', 'price': 360, 'dimensions': '43cm × 38cm × 28cm', 'tags': 'organic, modern', 'ai_pick': 'no'},
    'starberry':   {'name': 'Starberry', 'price': 0, 'dimensions': 'standard size', 'tags': 'playful, modern', 'ai_pick': 'no'},
    'strawberrys': {'name': 'Strawberry', 'price': 0, 'dimensions': 'standard size', 'tags': 'playful, modern', 'ai_pick': 'no'},
    'ripple-side-table-a-vanilla-noir':        {'name': 'Ripple Table A — Vanilla Noir', 'price': 535, 'dimensions': '36cm × 30cm × 57cm', 'tags': 'modern, minimal', 'ai_pick': 'yes'},
    'ripple-side-table-b-marble':              {'name': 'Ripple Table B — Marble', 'price': 360, 'dimensions': '36cm × 30cm × 57cm', 'tags': 'modern, minimal', 'ai_pick': 'yes'},
    'ripple-side-table-b-pomegranate-pattern': {'name': 'Ripple Table B — Pomegranate', 'price': 360, 'dimensions': '36cm × 30cm × 57cm', 'tags': 'modern, minimal', 'ai_pick': 'no'},
    'ripple-side-table-b-vanilla-noir':        {'name': 'Ripple Table B — Vanilla Noir', 'price': 360, 'dimensions': '36cm × 30cm × 57cm', 'tags': 'modern, minimal', 'ai_pick': 'yes'},
}


def prettify(s):
    return re.sub(r'\b\w', lambda m: m.group().upper(), s.replace('-', ' ').replace('_', ' '))


def scan_folders():
    """Return {id: category} from the catalog folder layout."""
    found = {}
    for path in glob.glob(os.path.join(CATALOG_DIR, '*', '*')):
        if not path.lower().endswith(IMG_EXT):
            continue
        category = os.path.basename(os.path.dirname(path))
        pid = os.path.splitext(os.path.basename(path))[0]
        found[pid] = category
    return found


def read_xlsx():
    """Return {id: {col: value}} from the Excel, if it exists."""
    rows = {}
    if not os.path.exists(XLSX):
        return rows
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        pid = (rec.get('id') or '').strip() if isinstance(rec.get('id'), str) else rec.get('id')
        if pid:
            rows[str(pid)] = rec
    return rows


def main():
    folders = scan_folders()
    xlsx = read_xlsx()

    meta = {}
    table = []  # rows to write back to the Excel
    for pid in sorted(folders, key=lambda p: (folders[p], p)):
        category = folders[pid]
        src = xlsx.get(pid) or SEED.get(pid) or {}
        name = (src.get('name') or '').strip() if isinstance(src.get('name'), str) else src.get('name')
        name = name or prettify(pid)
        price = src.get('price')
        try:
            price = int(float(price)) if price not in (None, '') else 0
        except (TypeError, ValueError):
            price = 0
        dimensions = (src.get('dimensions') or '').strip() if isinstance(src.get('dimensions'), str) else (src.get('dimensions') or '')
        dimensions = dimensions or 'standard size'
        raw_tags = src.get('tags') or ''
        tags = [t.strip() for t in str(raw_tags).split(',') if t.strip()]
        ai_pick = str(src.get('ai_pick') or '').strip().lower() in TRUTHY

        meta[pid] = {'name': name, 'price': price, 'dimensions': dimensions, 'tags': tags, 'aiPick': ai_pick}
        table.append([pid, category, name, price, dimensions, ', '.join(tags), 'yes' if ai_pick else 'no'])

    # write meta.json
    with open(META_JSON, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
        f.write('\n')

    # rewrite the Excel so it mirrors the live catalogue
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'products'
    ws.append([c.upper() for c in COLUMNS])
    for r in table:
        ws.append(r)
    # readable column widths
    for col, width in zip('ABCDEFG', (34, 12, 30, 8, 22, 22, 9)):
        ws.column_dimensions[col].width = width
    wb.save(XLSX)

    print(f'{len(meta)} products synced')
    print(f'  -> {os.path.relpath(META_JSON, ROOT)}')
    print(f'  -> {os.path.relpath(XLSX, ROOT)}')
    for pid, m in meta.items():
        flag = '' if m['price'] else '  <-- price missing'
        print(f"   {folders[pid]:9} {pid:42} EUR {m['price']:<5} {m['dimensions']}{flag}")


if __name__ == '__main__':
    main()
